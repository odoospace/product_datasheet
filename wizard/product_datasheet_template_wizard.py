# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from odoo.exceptions import UserError, AccessError
from datetime import datetime, date
from io import BytesIO
import openpyxl
import xlsxwriter
import string
import base64
import tempfile
import binascii
import zipfile
import os


class ProductDatasheetTemplateWizard(models.TransientModel):
    _name = 'product.datasheet.template.wizard'

    template_id = fields.Many2one('product.datasheet.template', string='Template')
    product_ids = fields.Many2many('product.product', string='Product')
    file_generated = fields.Binary(string='Generated ZIP file', attachment=False)

    @api.onchange('template_id')
    def onchange_template_id(self):
        return {'value': {'file_generated': False}}

    def action_download_excel(self):
        if not self.template_id:
            raise UserError(_("You don't select a Template!"))
        print(f'You have chosen {self.template_id.name}')
        tmp_url = '/tmp/ProductDatasheetTemplate/'
        try:
            os.stat(tmp_url)
        except:
            os.mkdir(tmp_url)

        zf = zipfile.ZipFile(tmp_url + 'ProductDatasheetTemplates.zip', mode='w')

        for product in self.product_ids:
            fp = tempfile.NamedTemporaryFile(suffix=".xlsx")
            fp_name = f'/tmp/{product.name}.xlsx'
            os.rename(fp.name, fp_name)
            fp.write(binascii.a2b_base64(self.template_id.file))  # self.xls_file is your binary field
            fp.seek(0)
            # path = '/home/file.xlsx'
            wb_obj = openpyxl.load_workbook(fp_name)
            sheet_obj = wb_obj.active
            sheet_obj.title = product.name
            general_dict = {
                'o.name': product.name,
                'o.image_1920': product.image_1920,
                'h.date': datetime.now().strftime('%Y/%m/%d'),
                'h.logo': self.env.user.company_id.logo,
                'h.regulation_footer': self.env['ir.config_parameter'].sudo().get_param(
                    'product_datasheet.regulation_footer'),
                'h.text_footer': self.env['ir.config_parameter'].sudo().get_param(
                    'product_datasheet.text_footer'),
            }
            for row in sheet_obj.iter_rows():
                for cell in row:
                    cell_value = cell.value
                    if cell_value is not None:
                        print(cell_value)
                        if cell_value.startswith('{{'):
                            cell_value = cell_value.replace(' ', '').replace('{{', '').replace('}}', '')
                            if 'i.' in cell_value:
                                if '|' in cell_value:
                                    label = cell_value.split('|')[1]
                                    cell_value_splitted = cell_value.split('|')[0].split('.')
                                    model_obj = cell_value_splitted[1]
                                    code = cell_value_splitted[2]
                                    model_env = f'product.datasheet.{model_obj}'
                                    if label == 'name':
                                        product_datasheet = self.env[model_env].search([('code', '=', code)])
                                        if product_datasheet:
                                            cell.value = product_datasheet.name
                                    else:
                                        section_code = cell_value_splitted[2]
                                        group_code = cell_value_splitted[3]
                                        field_code = cell_value_splitted[4]
                                        product_datasheet = self.env[model_env].search(
                                            [('product_id', '=', product.id),
                                             ('section_id.code', '=', section_code), ('group_id.code', '=', group_code),
                                             ('field_id.code', '=', field_code)])
                                        if product_datasheet:
                                            if product_datasheet.value:
                                                uom = ''
                                                uom_key = False
                                                if product_datasheet.field_id.uom_ids:
                                                    field_uom = product_datasheet.field_id.uom_ids.filtered(
                                                        lambda m: m.group_id.id == product_datasheet.group_id.id)
                                                    if field_uom:
                                                        uom_key = field_uom.uom
                                                else:
                                                    if product_datasheet.uom:
                                                        uom_key = product_datasheet.uom
                                                if uom_key:
                                                    uom = _(
                                                        dict(self.env[model_env].fields_get(
                                                            allfields=['uom'])['uom'][
                                                                 'selection'])[
                                                            uom_key])
                                                if product_datasheet.field_id and product_datasheet.field_id.type == 'number':
                                                    info_display = str(round(float(product_datasheet.value), 2))
                                                elif product_datasheet.field_id and product_datasheet.field_id.type == 'boolean':
                                                    if product_datasheet.value in ('0', 'True'):
                                                        info_display = _('Yes')
                                                    else:
                                                        info_display = _('No')
                                                else:
                                                    info_display = product_datasheet.value
                                                info_display += ' ' + uom if uom else ''
                                            else:
                                                info_display = '-'
                                            cell.value = info_display
                                else:
                                    cell_value_splitted = cell_value.split('.')
                                    model_obj = cell_value_splitted[1]
                                    code = cell_value_splitted[2]
                                    model_env = f'product.datasheet.{model_obj}'
                                    product_datasheet = self.env[model_env].search([('code', '=', code)])
                                    if product_datasheet:
                                        cell.value = product_datasheet.name

                                    # IMAGES SECTION
                                    if 'section' in cell_value:
                                        letter_column = list(string.ascii_uppercase)  # Array from A to Z
                                        columns_section = product_datasheet.column_ids.filtered(
                                            lambda m: m.group_id.export in [True])
                                        index_start = len(columns_section) + 1 if product_datasheet.column_ids else 2
                                        for product_image in product.image_ids.filtered(
                                                lambda m: m.section_id.id == product_datasheet.id):
                                            if product_image.image:
                                                buf_product_image = BytesIO(base64.b64decode(product_image.image))
                                                image_drawing = openpyxl.drawing.image.Image(buf_product_image)
                                                image_drawing.height = 300
                                                image_drawing.width = 300
                                                sheet_obj.add_image(image_drawing, letter_column[index_start] + str(
                                                    cell.row + 1))  # Insert section image
                                                index_start += 1
                            else:
                                if cell_value in general_dict and any(st in cell_value for st in ['image', 'logo']):
                                    if general_dict[cell_value]:
                                        buf_product_image = BytesIO(base64.b64decode(general_dict[cell_value]))
                                        image_drawing = openpyxl.drawing.image.Image(buf_product_image)
                                        image_drawing.height = 100
                                        image_drawing.width = 150
                                        sheet_obj.add_image(image_drawing, cell.coordinate)  # Insert product image
                                    cell.value = None
                                else:
                                    cell.value = general_dict[cell_value] if cell_value in general_dict else None
                        print(cell_value)

            wb_obj.save(fp_name)
            # output = BytesIO()
            # wb_obj.save(output)
            wb_obj.close()

            # output.seek(0)
            # data = output.read()

            # data = output.getvalue()
            # self.file_generated = base64.encodestring(data)

            zf.write(fp_name)

        zf.close()
        data = None
        with open(tmp_url + 'ProductDatasheetTemplates.zip', 'rb') as f:
            data = f.read()
        self.file_generated = base64.encodestring(data)

        return {
            'context': self.env.context,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'product.datasheet.template.wizard',
            'res_id': self.id,
            'view_id': False,
            'type': 'ir.actions.act_window',
            'target': 'new',
        }

    def action_download_layout(self):
        filename = f'Layout Delisano.xlsx'
        output = BytesIO()

        _info = {
            'code': 'Layout Delisano',
            'created': datetime.now().strftime('%Y/%m/%d')
        }

        workbook = xlsxwriter.Workbook(output)

        # TEXT FORMAT
        bold = workbook.add_format({'bold': True})
        italic = workbook.add_format({'italic': True})
        italic.set_font_size(10)

        # CELL FORMAT
        black_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': 'black'
        })
        normal_format = workbook.add_format({
            'font_color': 'black',
            'bg_color': 'white',
            'border': 1
        })

        # TAB NAME
        worksheet = workbook.add_worksheet(self.display_name)  # Tab with display_name of product

        letter_column = list(string.ascii_uppercase)  # Array from A to Z
        for letter in letter_column:  # Set width column from A to Z
            worksheet.set_column(letter + ':' + letter, 25)

        # FIELD -> SECTION AND GROUP IF FIELD IS DUPLICATED INFO PRODUCT
        fields_to_print = [
            'DEFAULT_CODE',
            'DCP',
            'DPUF',
            'NP',
            'DPE',
            'PA',
            'DEN',
            'CTA',
            'CDVC',
            'ING',
            'PRP',
            'VUC',
            'VUR',
            'CH',
            'CPA',
            'TEXTO',
            'CNP',
            'DPPAL',
            'DPPAN',
            'DPPL',
            'VE1-IN-VM100',
            'VE2-IN-VM100',
            'GR-IN-VM100',
            'GRS-IN-VM100',
            'CAH-IN-VM100',
            'CAHG-IN-VM100',
            'FA-IN-VM100',
            'PR-IN-VM100',
            'NA-IN-VM100',
            'K-IN-VM100',
            'FE-IN-VM100',
            'MG-IN-VM100',
            'CA-IN-VM100',
            'P-IN-VM100',
            'ZN-IN-VM100',
            'VB12-IN-VM100',
            'GLU-AOI-ALD',
            'GLU-AOI-ALI',
            'CRU-AOI-ALD',
            'CRU-AOI-ALI',
            'HUE-AOI-ALD',
            'HUE-AOI-ALI',
            'PES-AOI-ALD',
            'PES-AOI-ALI',
            'CAC-AOI-ALD',
            'CAC-AOI-ALI',
            'SOJ-AOI-ALD',
            'SOJ-AOI-ALI',
            'LEC-AOI-ALD',
            'LEC-AOI-ALI',
            'FRU-AOI-ALD',
            'FRU-AOI-ALI',
            'API-AOI-ALD',
            'API-AOI-ALI',
            'MOS-AOI-ALD',
            'MOS-AOI-ALI',
            'GRA-AOI-ALD',
            'GRA-AOI-ALI',
            'DIO-AOI-ALD',
            'DIO-AOI-ALI',
            'LUP-AOI-ALD',
            'LUP-AOI-ALI',
            'RAM-AM-M',
            'MOL-AOI-ALD',
            'MOL-AOI-ALI',
            'STA-AM-M',
            'ENT-AM-M',
            'LIS-AM-M',
            'BAC-AM-M',
            'LEV-AM-M',
            'MOH-AM-M',
            'SALM-AM-M',
            'BACA-AM-M',
            'ECOL-AM-M',
            'MDP1'
        ]

        pdinfo_obj = self.env['product.datasheet.info']
        pdsection_obj = self.env['product.datasheet.section']
        pdgroup_obj = self.env['product.datasheet.group']
        pdfield_obj = self.env['product.datasheet.field']

        i = 0
        j = 0
        for key in fields_to_print:
            if j == 0:
                worksheet.write(i, j, 'Default Code', black_format)
            else:
                key_edit = key.split('-')[0] if '-' in key else key
                pdfield = pdfield_obj.search([('code', '=', key_edit)])
                if pdfield:
                    name_column = ''
                    name_column += pdgroup_obj.search([('code', '=', 'ALD')],
                                                      limit=1).name + ' - ' if 'ALD' in key else ''
                    name_column += pdgroup_obj.search([('code', '=', 'ALI')],
                                                      limit=1).name + ' - ' if 'ALI' in key else ''
                    name_column += pdfield.name
                    name_column += ' (' + dict(pdfield._fields['uom'].selection).get(
                        pdfield.uom) + ')' if pdfield.uom else ''
                    worksheet.write(i, j, name_column, black_format)

            j += 1

        i += 1
        for product in self.product_ids:
            j = 0
            for key in fields_to_print:
                if j == 0:
                    worksheet.write(i, j, product.default_code, normal_format)
                else:
                    key_edit = key.split('-')[0] if '-' in key else key
                    pdfield = pdfield_obj.search([('code', '=', key_edit)])
                    if pdfield:
                        print(pdfield.code)
                        if '-' in key:
                            pdsection = pdsection_obj.search([('code', '=', key.split('-')[1])])
                            pdgroup = pdgroup_obj.search([('code', '=', key.split('-')[2])])
                            if pdsection and pdgroup:
                                pdinfo = pdinfo_obj.search([('product_id', '=', product.id),
                                                            ('section_id', '=', pdsection.id),
                                                            ('group_id', '=', pdgroup.id),
                                                            ('field_id', '=', pdfield.id)])
                        else:
                            pdinfo = pdinfo_obj.search([('product_id', '=', product.id),
                                                        ('field_id', '=', pdfield.id)])

                        worksheet.write(i, j, pdinfo.value if pdinfo else '', normal_format)
                j += 1
            print(product.name)
            i += 1

        print('Saving excel...')
        workbook.close()
        output.seek(0)

        data = output.read()
        self.file_generated = base64.encodestring(data)

        return {
            'context': self.env.context,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'product.datasheet.template.wizard',
            'res_id': self.id,
            'view_id': False,
            'type': 'ir.actions.act_window',
            'target': 'new',
        }
