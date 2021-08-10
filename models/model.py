from datetime import datetime, date

from odoo import models, fields, api, _
from io import BytesIO
import shortuuid
import xlsxwriter
import openpyxl
import base64
import json
import string
import html2text


class ResConfigSettings(models.TransientModel):
    _inherit = 'res.config.settings'

    regulation_footer = fields.Text('Regulations', help='Regulations in Footer of Excel', translate=True)
    text_footer = fields.Text('Text Footer', help='Text in Footer of Excel', translate=True)

    @api.model
    def get_values(self):
        res = super(ResConfigSettings, self).get_values()
        res.update(
            regulation_footer=self.env[
                'ir.config_parameter'].sudo().get_param(
                'product_datasheet.regulation_footer'),
            text_footer=self.env[
                'ir.config_parameter'].sudo().get_param(
                'product_datasheet.text_footer'),
        )
        return res

    def set_values(self):
        super(ResConfigSettings, self).set_values()
        self.env['ir.config_parameter'].sudo().set_param(
            'product_datasheet.regulation_footer',
            self.regulation_footer)
        self.env['ir.config_parameter'].sudo().set_param(
            'product_datasheet.text_footer',
            self.text_footer)


# TODO: write historic info
class Section(models.Model):
    _name = 'product.datasheet.section'
    _description = "Product Datasheet Section"

    code = fields.Char(required=True)
    name = fields.Char(required=True, translate=True)
    active = fields.Boolean(default=True)
    timestamp = fields.Datetime(default=fields.Datetime.now)
    export = fields.Boolean('Is it exported?')

    group_ids = fields.One2many('product.datasheet.group', 'section_id')


class Group(models.Model):
    _name = 'product.datasheet.group'
    _description = "Product Datasheet Group"

    code = fields.Char(required=True)
    name = fields.Char(required=True, translate=True)
    timestamp = fields.Datetime(default=fields.Datetime.now)
    active = fields.Boolean(default=True)
    export = fields.Boolean('Is it exported?')

    section_id = fields.Many2one('product.datasheet.section')

    def name_get(self):
        result = []
        for group in self:
            name = f'{group.name} ({group.section_id.name})'
            result.append((group.id, name))
        return result


class Field(models.Model):
    _name = "product.datasheet.field"
    _description = "Product Datasheet Field"

    @api.depends('info_ids')
    def _compute_section_group(self):
        for record in self:
            record.section_id = record.info_ids[-1].section_id.id if record.info_ids and record.info_ids[
                -1].section_id else False
            record.group_id = record.info_ids[-1].group_id.id if record.info_ids and record.info_ids[
                -1].group_id else False

    code = fields.Char(required=True)
    name = fields.Char(required=True, translate=True)
    type = fields.Selection(
        [
            ("integer", "Integer"),
            ("string", "String"),
            ("html", "HTML"),
            ("selection", "Selection"),  # comma separated values or so
        ], required=True, translate=True)
    uom = fields.Selection(
        [
            ("gr", _("gr")),
            ("cfu_g", _("cfu/gr")),
            ("m3", _("m³")),
            ("cm", _("cm")),
            ("cm3", _("cm³")),
            ("mm", _("mm")),
            ("µg", _("µg")),
            ("box", _("caja")),
            ("mg", _("mg")),
            ("kcal", _("kcal")),
            ("KJ", _("kj")),
            ("ud", _("unidades")),
            ("kg", _("kg")),
            ("l", _("l")),
            ("min", _("min")),
            ("seg", _("seg")),
            ("day", _("day")),
            ("month", _("month")),
        ])
    export = fields.Boolean('Is it exported?')

    info_ids = fields.One2many('product.datasheet.info', 'field_id')
    section_id = fields.Many2one('product.datasheet.section', compute=_compute_section_group, store=True)
    group_id = fields.Many2one('product.datasheet.group', compute=_compute_section_group, store=True)


class Info(models.Model):
    _name = 'product.datasheet.info'
    _description = 'Product Datasheet Info'
    _order = 'sequence'

    @api.depends('value')
    def _compute_value_name(self):
        for record in self:
            # This will be called every time the value field changes
            if record.value and len(record.value) > 50:
                record.value_display = record.value[:47] + '...'
            else:
                record.value_display = record.value

    field_id = fields.Many2one('product.datasheet.field', required=True)
    value = fields.Text(translate=True)
    value_display = fields.Text(compute=_compute_value_name)
    timestamp = fields.Datetime(default=fields.Datetime.now)
    active = fields.Boolean(default=True)

    product_id = fields.Many2one('product.product')
    group_id = fields.Many2one('product.datasheet.group', required=True)
    # related fields
    group_name = fields.Char(string=_('Group'), related='group_id.name')
    section_id = fields.Many2one(related='group_id.section_id')
    uom = fields.Selection(related='field_id.uom')
    sequence = fields.Integer(default=1)


class ProductDatasheetImage(models.Model):
    _name = 'product.datasheet.image'
    _description = 'Product Datasheet Image'
    _order = 'section_id'

    section_id = fields.Many2one('product.datasheet.section')
    image = fields.Binary(string='Image file', required=True, attachment=False)
    product_id = fields.Many2one('product.product')


class ProductProduct(models.Model):
    _inherit = 'product.product'

    def filter_by_name(self):
        res = []
        if self.filter_field:
            res = [('field_id.name', 'ilike', self.filter_field)]
        return res

    info_ids = fields.One2many('product.datasheet.info', 'product_id', domain=filter_by_name)
    image_ids = fields.One2many('product.datasheet.image', 'product_id')

    datasheet_note = fields.Text()
    country_ids = fields.Many2many('res.country', 'product_ids')

    # filters
    filter_field = fields.Char('Field')
    filter_section = fields.Many2one('product.datasheet.section')
    filter_group = fields.Many2one('product.datasheet.group')

    # add the field itself to onchange to trigger this method in edit mode too
    @api.onchange('filter_field')
    def onchange_filter_field(self):
        print('***')
        domain = []
        if self.filter_field:
            domain.append(('field_id.name', 'ilike', self.filter_field))
        if self.filter_section:
            domain.append(('field_id.section_id', 'ilike', self.filter_section))
        if self.filter_group:
            domain.append(('field_id.group_id', 'ilike', self.filter_group))
        res = {'domain': {'info_ids': domain}}
        print(res)
        self.info_ids = []
        return res

    @api.model
    def duplicate_product(self):
        for product_product in self:
            # COPY PRODUCT TEMPLATE TO GENERATE PRODUCT VARIANT
            product_template_copy = product_product.product_tmpl_id.copy()
            product_product_copy = product_template_copy.product_variant_id

            # DATASHEET INFO TAB
            filter_field = product_product.filter_field
            filter_section = product_product.filter_section.id if product_product.filter_section else False
            filter_group = product_product.filter_group.id if product_product.filter_group else False
            product_product_copy.country_ids = [(6, 0, product_product.country_ids.ids)]
            product_product_copy.datasheet_note = product_product.datasheet_note
            product_product_copy.filter_field = filter_field
            product_product_copy.filter_section = filter_section
            product_product_copy.filter_group = filter_group

            product_datasheet_info = self.env['product.datasheet.info']
            for info in product_datasheet_info.search([('product_id', '=', product_product.id)]):
                product_datasheet_info.create({
                    'sequence': info.sequence,
                    'section_id': info.section_id.id if info.section_id else False,
                    'group_id': info.group_id.id if info.group_id else False,
                    'field_id': info.field_id.id if info.field_id else False,
                    'value_display': info.value_display,
                    'uom': info.uom,
                    'product_id': product_product_copy.id,
                })

            product_datasheet_image = self.env['product.datasheet.image']
            for image in product_product.image_ids:
                product_datasheet_image.create({
                    'section_id': image.section_id.id if image.section_id else False,
                    'image': image.image,
                    'product_id': product_product_copy.id,
                })

            # GENERAL INFO TAB
            product_product_copy.standard_price = product_product.standard_price
            product_secondary_unit = self.env['product.secondary.unit']
            for secondary_uom in product_product.secondary_uom_ids:
                product_secondary_unit.create({
                    'code': secondary_uom.code,
                    'name': secondary_uom.name,
                    'factor': secondary_uom.factor,
                    'uom_id': secondary_uom.uom_id.id if secondary_uom.uom_id else False,
                    'product_tmpl_id': product_template_copy.id,
                })

            # PURCHASE INFO TAB
            product_supplierinfo = self.env['product.supplierinfo']
            for seller in product_product.seller_ids:
                product_supplierinfo.create({
                    'name': seller.name.id if seller.name else False,
                    'product_name': seller.product_name,
                    'product_code': seller.product_code,
                    'x_studio_mtodo_de_transporte_1': seller.x_studio_mtodo_de_transporte_1,
                    'x_studio_incoterm_1': seller.x_studio_incoterm_1,
                    'delay': seller.delay,
                    'min_qty': seller.min_qty,
                    'product_uom': seller.product_uom.id if seller.product_uom else False,
                    'price': seller.price,
                    'currency_id': seller.currency_id.id if seller.currency_id else False,
                    'date_start': seller.date_start,
                    'date_end': seller.date_end,
                    'product_id': product_product_copy.id if seller.product_id else False,
                    'product_tmpl_id': product_template_copy.id,
                })

    @api.model
    def change_sequence_datasheet(self):
        print('START!')
        product_datasheet_info = self.env['product.datasheet.info']
        for product_product in self:
            print(product_product.name)
            products = self.env['product.product'].search([('id', '!=', product_product.id)])
            print(str(len(products)))
            for product_to_edit in products:
                print(f'Product ID: {str(product_to_edit.id)}')
                for info_product_product in product_product.info_ids:
                    info_product_to_edit = product_datasheet_info.search([
                        ('product_id', '=', product_to_edit.id),
                        ('section_id', '=', info_product_product.section_id.id),
                        ('group_id', '=', info_product_product.group_id.id),
                        ('field_id', '=', info_product_product.field_id.id)
                    ])
                    if info_product_to_edit:
                        print(f'Section: {str(info_product_to_edit.section_id.name)}, '
                              f'Group: {str(info_product_to_edit.group_id.name)}, '
                              f'Field: {str(info_product_to_edit.field_id.name)}, '
                              f'Value: {str(info_product_to_edit.value_display)}')
                        info_product_to_edit.sequence = info_product_product.sequence
        print('END!')

    def download_xlsx(self):
        # TODO: reload page to refresh attachments
        filename = f'{self.name}.xlsx'
        output = BytesIO()

        _info = {
            'code': 'DataSheet of Product',
            'created': datetime.now().strftime('%Y/%m/%d')
        }

        workbook = xlsxwriter.Workbook(output)

        # TEXT FORMAT
        bold = workbook.add_format({'bold': True})
        italic = workbook.add_format({'italic': True})
        italic.set_font_size(10)
        red = workbook.add_format({'color': 'red'})
        blue = workbook.add_format({'color': 'blue'})
        center = workbook.add_format({'align': 'center'})
        superscript = workbook.add_format({'font_script': 1})

        # CELL FORMAT
        product_name_format = workbook.add_format({
            'bold': True,
            'font_color': 'black',
            'bg_color': 'white',
            'border': 1
        })
        product_name_format.set_font_size(20)
        product_name_format.set_align('center')
        product_name_format.set_align('vcenter')
        black_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': 'black'
        })
        gray_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': 'gray'
        })
        normal_format = workbook.add_format({
            'font_color': 'black',
            'bg_color': 'white',
            'border': 1
        })
        normal_center_format = workbook.add_format({
            'font_color': 'black',
            'bg_color': 'white',
            'border': 1
        })
        normal_center_format.set_align('center')
        normal_center_format.set_align('vcenter')
        footer_format = workbook.add_format({
            'font_color': 'black',
            'bg_color': 'white'
        })
        footer_format.set_font_size(10)

        # TAB NAME
        worksheet = workbook.add_worksheet(self.display_name)  # Tab with display_name of product

        # COMMENTS
        # info = _info
        # code = shortuuid.uuid()
        # info['worksheet'] = code
        #
        # worksheet.write_comment('A1', json.dumps(info))

        # INFO COMPANY HEADER
        if self.env.user.company_id and self.env.user.company_id.logo:
            buf_image_company = BytesIO(base64.b64decode(self.env.user.company_id.logo))
            worksheet.insert_image('A1', "image_company.png", {
                'image_data': buf_image_company,
                'x_scale': 0.03,
                'y_scale': 0.03
            })

        worksheet.set_row(0, 70)  # Set height of first row
        worksheet.set_column('A:A', 100)  # Set width column A
        worksheet.set_column('B:B', 50)  # Set width column B
        worksheet.set_column('C:C', 50)  # Set width column C
        letter_column = list(string.ascii_uppercase)  # Array from A to Z
        for letter in letter_column[3:]:  # Set width column from D to Z
            worksheet.set_column(letter + ':' + letter, 25)

        worksheet.write(0, 0, self.name, product_name_format)
        worksheet.write(0, 1, datetime.now().strftime('%Y/%m/%d'), normal_center_format)

        # DATA OF SUPPLIER
        if self._context['lang'] == 'es_ES':
            title_data_supplier = ['Datos del Proveedor', 'Nombre Empresa', 'CIF', 'Registro Sanitario',
                                   'Dirección Fiscal', 'Contacto', 'Página Web']
        else:
            title_data_supplier = ['Supplier Data', 'Company Name', 'CIF', 'Health Register',
                                   'Fiscal Address', 'Contact', 'Website']

        row_start = 2
        row_title_supplier = row_start
        row_data_supplier = row_start + 1

        worksheet.write(row_title_supplier, 1, '', black_format)

        for title in title_data_supplier:
            if row_title_supplier == 2:
                format_title = black_format
            else:
                format_title = normal_format
            worksheet.write(row_title_supplier, 0, title, format_title)
            row_title_supplier += 1

        foodsfortomorrow_company = self.env['res.company'].sudo().search([('id', '=', 1)])
        if foodsfortomorrow_company:
            direction = foodsfortomorrow_company.street + ' - ' + foodsfortomorrow_company.zip + ', ' + foodsfortomorrow_company.state_id.display_name
            data_supplier = [foodsfortomorrow_company.name, foodsfortomorrow_company.vat,
                             foodsfortomorrow_company.company_registry, direction,
                             'calidad@heurafoods.com', foodsfortomorrow_company.website]
            for data in data_supplier:
                worksheet.write(row_data_supplier, 1, data, normal_format)
                row_data_supplier += 1

        # IMAGE PRODUCT
        if self.image_1920:
            row_data_product = row_start + 1
            buf_image_product = BytesIO(base64.b64decode(self.image_1920))
            worksheet.insert_image('C' + str(row_data_product), "image_product.png", {
                'image_data': buf_image_product,
                'x_scale': 0.3,
                'y_scale': 0.3
            })  # Insert image product

        # DATA OF PRODUCT
        row_start = row_title_supplier + 1  # Space between tables

        for section in self.env['product.datasheet.section'].search([('export', '=', True)]):
            worksheet.write(row_start, 0, section.name, black_format)
            worksheet.write(row_start, 1, '', black_format)
            row_start += 1
            # IMAGES SECTION
            for product_image in self.image_ids.filtered(lambda m: m.section_id.id == section.id):
                if product_image.image:
                    buf_product_image = BytesIO(base64.b64decode(product_image.image))
                    worksheet.insert_image(letter_column[2] + str(row_start + 1),
                                           "product_image.png", {
                                               'image_data': buf_product_image,
                                               'x_scale': 0.3,
                                               'y_scale': 0.3
                                           })  # Insert product image
            for group in section.group_ids.filtered(lambda m: m.export in [True]):
                worksheet.write(row_start, 0, group.name, gray_format)
                worksheet.write(row_start, 1, '', gray_format)
                row_start += 1
                for info in self.env['product.datasheet.info'].search(
                        [('product_id', '=', self.id), ('section_id', '=', section.id), ('group_id', '=', group.id)],
                        order='sequence'):

                    def isfloat(value):
                        try:
                            float(value)
                            return True
                        except ValueError:
                            return False

                    # GET VALUE DISPLAY
                    if info.field_id and info.field_id.export:
                        if info.value and info.value != 'False':
                            uom = ''
                            if info.uom:
                                uom = _(
                                    dict(self.env['product.datasheet.info'].fields_get(allfields=['uom'])['uom'][
                                             'selection'])[
                                        info.uom])
                            if isfloat(info.value):
                                info_display = str(round(float(info.value), 2))
                            else:
                                info_display = info.value
                            info_display += ' ' + uom if uom else ''
                        else:
                            info_display = '-'

                        worksheet.write(row_start, 0, info.field_id.name, normal_format)
                        worksheet.write(row_start, 1, info_display, normal_format)
                        row_start += 1

        # FOOTER
        regulation_footer = self.env['ir.config_parameter'].sudo().get_param('product_datasheet.regulation_footer')
        text_footer = self.env['ir.config_parameter'].sudo().get_param('product_datasheet.text_footer')

        worksheet.set_row(row_start + 3, 250)  # Set height of row
        worksheet.write(row_start + 3, 0, regulation_footer, footer_format)

        worksheet.set_row(row_start + 4, 70)  # Set height of row
        text_footer_splitted = text_footer.split('\n') if text_footer and '\n' in text_footer else ''
        if len(text_footer_splitted) == 3:
            worksheet.write_rich_string('A' + str(row_start + 5),
                                        bold, text_footer_splitted[0] + '\n',
                                        italic,
                                        text_footer_splitted[1] + '\n',
                                        text_footer_splitted[2] + '\n')

        print('Saving excel...')
        workbook.close()
        output.seek(0)

        data = output.read()
        attachment = self.add_file_in_attachment(filename, data)
        # TODO: refresh page
        return {
            'type': 'ir.actions.act_url',
            'url': "web/content/?model=ir.attachment&id=" + str(attachment.id) +
                   f"&filename={self.name}.xlsx&field=datas&download=true&filename=" + attachment.name,
            'target': 'new',
        }

    def generate_template_xlsx(self):
        # TODO: reload page to refresh attachments
        filename = f'{self.name}.xlsx'
        output = BytesIO()

        _info = {
            'code': 'DataSheet of Product',
            'created': datetime.now().strftime('%Y/%m/%d')
        }

        workbook = xlsxwriter.Workbook(output)

        # TEXT FORMAT
        bold = workbook.add_format({'bold': True})
        italic = workbook.add_format({'italic': True})
        italic.set_font_size(10)
        red = workbook.add_format({'color': 'red'})
        blue = workbook.add_format({'color': 'blue'})
        center = workbook.add_format({'align': 'center'})
        superscript = workbook.add_format({'font_script': 1})

        # CELL FORMAT
        product_name_format = workbook.add_format({
            'bold': True,
            'font_color': 'black',
            'bg_color': 'white',
            'border': 1
        })
        product_name_format.set_font_size(20)
        product_name_format.set_align('center')
        product_name_format.set_align('vcenter')
        black_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': 'black'
        })
        gray_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': 'gray'
        })
        normal_format = workbook.add_format({
            'font_color': 'black',
            'bg_color': 'white',
            'border': 1
        })
        normal_center_format = workbook.add_format({
            'font_color': 'black',
            'bg_color': 'white',
            'border': 1
        })
        normal_center_format.set_align('center')
        normal_center_format.set_align('vcenter')
        footer_format = workbook.add_format({
            'font_color': 'black',
            'bg_color': 'white'
        })
        footer_format.set_font_size(10)

        # TAB NAME
        worksheet = workbook.add_worksheet(f'{{{{ o.display_name }}}}')  # Tab with display_name of product

        # COMMENTS
        # info = _info
        # code = shortuuid.uuid()
        # info['worksheet'] = code
        #
        # worksheet.write_comment('A1', json.dumps(info))

        # INFO COMPANY HEADER
        if self.env.user.company_id and self.env.user.company_id.logo:
            buf_image_company = BytesIO(base64.b64decode(self.env.user.company_id.logo))
            worksheet.insert_image('A1', "image_company.png", {
                'image_data': buf_image_company,
                'x_scale': 0.03,
                'y_scale': 0.03
            })

        worksheet.set_row(0, 70)  # Set height of first row
        worksheet.set_column('A:A', 100)  # Set width column A
        worksheet.set_column('B:B', 50)  # Set width column B
        worksheet.set_column('C:C', 50)  # Set width column C
        letter_column = list(string.ascii_uppercase)  # Array from A to Z
        for letter in letter_column[3:]:  # Set width column from D to Z
            worksheet.set_column(letter + ':' + letter, 25)

        worksheet.write(0, 0, f'{{{{ o.name }}}}', product_name_format)
        worksheet.write(0, 1, f'{{{{ h.date }}}}', normal_center_format)

        # DATA OF SUPPLIER
        if self._context['lang'] == 'es_ES':
            title_data_supplier = ['Datos del Proveedor', 'Nombre Empresa', 'CIF', 'Registro Sanitario',
                                   'Dirección Fiscal', 'Contacto', 'Página Web']
        else:
            title_data_supplier = ['Supplier Data', 'Company Name', 'CIF', 'Health Register',
                                   'Fiscal Address', 'Contact', 'Website']

        row_start = 2
        row_title_supplier = row_start
        row_data_supplier = row_start + 1

        worksheet.write(row_title_supplier, 1, '', black_format)

        for title in title_data_supplier:
            if row_title_supplier == 2:
                format_title = black_format
            else:
                format_title = normal_format
            worksheet.write(row_title_supplier, 0, title, format_title)
            row_title_supplier += 1

        foodsfortomorrow_company = self.env['res.company'].sudo().search([('id', '=', 1)])
        if foodsfortomorrow_company:
            direction = foodsfortomorrow_company.street + ' - ' + foodsfortomorrow_company.zip + ', ' + foodsfortomorrow_company.state_id.display_name
            data_supplier = [foodsfortomorrow_company.name, foodsfortomorrow_company.vat,
                             foodsfortomorrow_company.company_registry, direction,
                             'calidad@heurafoods.com', foodsfortomorrow_company.website]
            for data in data_supplier:
                worksheet.write(row_data_supplier, 1, data, normal_format)
                row_data_supplier += 1

        # IMAGE PRODUCT
        if self.image_1920:
            row_data_product = row_start + 1
            buf_image_product = BytesIO(base64.b64decode(self.image_1920))
            worksheet.insert_image('C' + str(row_data_product), "image_product.png", {
                'image_data': buf_image_product,
                'x_scale': 0.3,
                'y_scale': 0.3
            })  # Insert image product

        # DATA OF PRODUCT
        row_start = row_title_supplier + 1  # Space between tables

        for section in self.env['product.datasheet.section'].search([('export', '=', True)]):
            worksheet.write(row_start, 0, f'{{{{ i.section.{section.code} }}}}', black_format)
            worksheet.write(row_start, 1, '', black_format)
            row_start += 1
            # IMAGES SECTION
            for product_image in self.image_ids.filtered(lambda m: m.section_id.id == section.id):
                if product_image.image:
                    buf_product_image = BytesIO(base64.b64decode(product_image.image))
                    worksheet.insert_image(letter_column[2] + str(row_start + 1),
                                           "product_image.png", {
                                               'image_data': buf_product_image,
                                               'x_scale': 0.3,
                                               'y_scale': 0.3
                                           })  # Insert product image
            for group in section.group_ids.filtered(lambda m: m.export in [True]):
                worksheet.write(row_start, 0, f'{{{{ i.group.{group.code} }}}}', gray_format)
                worksheet.write(row_start, 1, '', gray_format)
                row_start += 1
                for info in self.env['product.datasheet.info'].search(
                        [('product_id', '=', self.id), ('section_id', '=', section.id), ('group_id', '=', group.id)],
                        order='sequence'):
                    if info.field_id and info.field_id.export:
                        worksheet.write(row_start, 0, f'{{{{ i.field.{info.field_id.code} | name }}}}', normal_format)
                        worksheet.write(row_start, 1, f'{{{{ i.field.{info.field_id.code} | value }}}}', normal_format)
                        row_start += 1

        # FOOTER
        worksheet.set_row(row_start + 3, 250)  # Set height of row
        worksheet.write(row_start + 3, 0, f'{{{{ regulation_footer }}}}', footer_format)

        worksheet.set_row(row_start + 4, 70)  # Set height of row
        worksheet.write(row_start + 5, 0, f'{{{{ text_footer }}}}', footer_format)

        print('Saving excel...')
        workbook.close()
        output.seek(0)

        data = output.read()
        attachment = self.add_file_in_attachment(filename, data)
        # TODO: refresh page
        return {
            'type': 'ir.actions.act_url',
            'url': "web/content/?model=ir.attachment&id=" + str(attachment.id) +
                   f"&filename={self.name}.xlsx&field=datas&download=true&filename=" + attachment.name,
            'target': 'new',
        }

    def read_xlsx(self):
        path = '/home/file.xlsx'
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        cell_obj = sheet_obj.cell(row=1, column=1)
        print(cell_obj.value)
        cell_obj.value = 'New text'
        print(cell_obj.value)
        wb_obj.save('home/file.xlsx')

    def add_file_in_attachment(self, filename, output):
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': base64.b64encode(output),
            'db_datas': filename,
            'res_model': 'product.product',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        return attachment


class Country(models.Model):
    _inherit = 'res.country'

    product_ids = fields.Many2many('product.product', 'country_ids')
